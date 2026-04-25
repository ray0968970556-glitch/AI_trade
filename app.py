"""
📊 價量明細 + CDP 交易指標分析系統 (Gradio 版)
部署至 Hugging Face Space
"""

import subprocess, sys, os

# ── 安裝所需套件（Colab / HuggingFace 首次執行）─────────────────────────────
subprocess.run([
    sys.executable, "-m", "pip", "install", "-q",
    "gradio",
    "undetected-chromedriver",
    "selenium",
    "fake-useragent",
    "beautifulsoup4",
    "openpyxl",
], check=True)

# ── 安裝 Chrome / ChromeDriver（在 HuggingFace Space 終端機執行）──────────
def install_chrome():
    """安裝 Chrome 146 + ChromeDriver，僅需執行一次"""
    # 下載 ChromeDriver 146
    subprocess.run([
        "wget", "-q",
        "https://storage.googleapis.com/chrome-for-testing-public/146.0.7680.177/linux64/chromedriver-linux64.zip",
        "-O", "/tmp/chromedriver.zip"
    ], check=True)
    subprocess.run(["unzip", "-o", "/tmp/chromedriver.zip", "-d", "/usr/local/bin/"], check=True)
    subprocess.run(["mv", "/usr/local/bin/chromedriver-linux64/chromedriver", "/usr/local/bin/chromedriver"], check=False)
    subprocess.run(["chmod", "+x", "/usr/local/bin/chromedriver"], check=True)

    # 安裝 Google Chrome stable（若尚未安裝）
    if not os.path.exists("/usr/bin/google-chrome"):
        subprocess.run([
            "wget", "-q",
            "https://dl.google.com/linux/direct/google-chrome-stable_current_amd64.deb",
            "-O", "/tmp/chrome.deb"
        ], check=True)
        subprocess.run(["dpkg", "-i", "/tmp/chrome.deb"], check=False)
        subprocess.run(["apt-get", "install", "-f", "-y"], check=False)

# 啟動時自動安裝
try:
    install_chrome()
    print("✅ Chrome + ChromeDriver 安裝完成")
except Exception as e:
    print(f"⚠️ Chrome 安裝可能已完成或發生非致命錯誤: {e}")

# ── 套件 ────────────────────────────────────────────────────────────────────
import time, io
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import undetected_chromedriver as uc
from selenium.webdriver.chrome.options import Options
from fake_useragent import UserAgent
import gradio as gr

# ── 爬取 PChome 逐筆成交 ────────────────────────────────────────────────────
def fetch_pchome(stock_id: str) -> pd.DataFrame:
    url = f"https://pchome.megatime.com.tw/stock/sto0/ock3/sid{stock_id}.html"
    ua = UserAgent()
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument(f"user-agent={ua.random}")
    options.binary_location = "/usr/bin/google-chrome"

    driver = uc.Chrome(
        options=options,
        version_main=146,
        driver_executable_path="/usr/local/bin/chromedriver"
    )
    driver.get(url)
    time.sleep(5)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    data = []
    for tbody in soup.find_all("tbody"):
        for row in tbody.find_all("tr"):
            cols = [col.get_text(strip=True) for col in row.find_all("td")]
            if len(cols) >= 7 and ":" in cols[0]:
                data.append(cols[:7])

    if not data:
        raise ValueError(f"❌ 找不到股票代號 {stock_id} 的逐筆資料，請確認代號是否正確")

    df = pd.DataFrame(data, columns=["時間","買價","賣價","成交價","漲跌","分量(張)","累計量(張)"])
    # 清除非數字列
    df = df[df["時間"].str.contains(":")].reset_index(drop=True)
    for col in ["買價","賣價","成交價","漲跌","分量(張)","累計量(張)"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["成交價"]).reset_index(drop=True)
    return df

# ── CDP 計算 ─────────────────────────────────────────────────────────────────
def calc_cdp(h, l, c):
    cdp = (h + l + 2 * c) / 4
    ah  = cdp + (h - l)
    nh  = 2 * cdp - l
    nl  = 2 * cdp - h
    al  = cdp - (h - l)
    return cdp, ah, nh, nl, al

def zone_label(price, cdp, ah, nh, nl, al):
    if price > ah:    return '超強勢區'
    elif price > nh:  return '偏多區'
    elif price > cdp: return '多方中性'
    elif price > nl:  return '空方中性'
    elif price > al:  return '偏空區'
    else:             return '超弱勢區'

def get_zone_signal(price, cdp, ah, nh, nl, al):
    zone = zone_label(price, cdp, ah, nh, nl, al)
    signals = {
        '超強勢區': '❌ 謹慎追高，等回落',
        '偏多區':   '📈 多方格局，注意反轉',
        '多方中性': '➡️ 持多觀察',
        '空方中性': '➡️ 持空觀察',
        '偏空區':   '📉 空方格局，注意反轉',
        '超弱勢區': '❌ 謹慎追空，等反彈',
    }
    return zone, signals[zone]

# ── 30分K + 滾動CDP ──────────────────────────────────────────────────────────
def build_30min_kdp(df, prev_close):
    df2 = df.copy()
    df2['時間_dt'] = pd.to_datetime(df2['時間'], format='%H:%M:%S', errors='coerce')
    df2 = df2.dropna(subset=['時間_dt'])
    df_30 = df2.set_index('時間_dt').resample('30min')['成交價'].agg(
        開盤='first', 最高='max', 最低='min', 收盤='last'
    ).dropna()
    df_30['成交量(張)'] = df2.set_index('時間_dt').resample('30min')['分量(張)'].sum()
    df_30 = df_30.reset_index()
    df_30['時間'] = df_30['時間_dt'].dt.strftime('%H:%M')

    df_30[['CDP','AH','NH','NL','AL']] = np.nan
    for i in range(1, len(df_30)):
        ph, pl, pc = df_30.loc[i-1, '最高'], df_30.loc[i-1, '最低'], df_30.loc[i-1, '收盤']
        df_30.loc[i, ['CDP','AH','NH','NL','AL']] = calc_cdp(ph, pl, pc)

    # 第一根用昨收估算
    est_h = prev_close * 1.01
    est_l = prev_close * 0.99
    df_30.loc[0, ['CDP','AH','NH','NL','AL']] = calc_cdp(est_h, est_l, prev_close)

    df_30['區間'] = ''
    df_30['交易建議'] = ''
    for i, row in df_30.iterrows():
        if pd.notna(row['CDP']):
            z, s = get_zone_signal(row['收盤'], row['CDP'], row['AH'], row['NH'], row['NL'], row['AL'])
            df_30.loc[i, '區間'] = z
            df_30.loc[i, '交易建議'] = s

    df_30['進場信號'] = ''
    for i in range(1, len(df_30)):
        prev_c = df_30.loc[i-1, '收盤']
        curr_c = df_30.loc[i, '收盤']
        nh_, nl_, ah_, al_ = df_30.loc[i, 'NH'], df_30.loc[i, 'NL'], df_30.loc[i, 'AH'], df_30.loc[i, 'AL']
        sigs = []
        if pd.notna(nh_):
            if prev_c < nh_ <= curr_c:  sigs.append('🔺 突破NH → 做多')
            if prev_c > nh_ >= curr_c:  sigs.append('🔻 跌破NH → 做空/停利')
            if prev_c < nl_ <= curr_c:  sigs.append('🔺 突破NL → 逢低買進')
            if prev_c > al_ >= curr_c:  sigs.append('🔻 跌破AL → 強勢做空')
            if prev_c < ah_ <= curr_c:  sigs.append('🔺 突破AH → 強勢突破')
        df_30.loc[i, '進場信號'] = ' / '.join(sigs) if sigs else '—'

    return df_30

# ── 輸出美化 Excel ────────────────────────────────────────────────────────────
def build_excel(df, df_30, cdp_vals, ohlcv):
    CDP, AH, NH, NL, AL = cdp_vals
    O, H, L, C, total_vol = ohlcv

    HEADER_BG, HEADER_FG = '1F3A6E', 'FFFFFF'
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    zone_bg = {
        '超強勢區': 'FFD7D7', '偏多區': 'FFEAD7', '多方中性': 'D7F5D7',
        '空方中性': 'D6E4F0', '偏空區': 'E8D6F0', '超弱勢區': 'FFD0D0',
    }
    zone_fc = {
        '超強勢區': 'FF4500', '偏多區': 'FF8C00', '多方中性': '2E8B57',
        '空方中性': '4682B4', '偏空區': '6A5ACD', '超弱勢區': 'DC143C',
    }

    def style_header(ws, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=2, column=c)
            cell.font = Font(bold=True, color=HEADER_FG, name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[2].height = 22

    def add_title(ws, title, ncols):
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = Font(bold=True, size=12, color='FFFFFF', name='Arial')
        tc.fill = PatternFill('solid', start_color='1F3A6E')
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

    def auto_width(ws):
        for col in ws.columns:
            mx = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 30)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Sheet1
        df_out = df[['時間','買價','賣價','成交價','漲跌','分量(張)','累計量(張)','CDP區間']].copy()
        df_out.to_excel(writer, sheet_name='逐筆明細', index=False)
        # Sheet2
        col_order = ['時間','開盤','最高','最低','收盤','成交量(張)',
                     'CDP','AH','NH','NL','AL','區間','進場信號','交易建議']
        df_30_out = df_30[col_order].copy()
        for c in ['CDP','AH','NH','NL','AL']:
            df_30_out[c] = df_30_out[c].round(2)
        df_30_out.to_excel(writer, sheet_name='30分K線CDP分析', index=False)
        # Sheet3 Summary
        summary = pd.DataFrame({
            '指標': ['開盤(O)','最高(H)','最低(L)','收盤(C)','─────',
                     'CDP（核心軸線）','AH（最高目標）','NH（偏多壓力）','NL（偏空支撐）','AL（最低目標）'],
            '數值': [O,H,L,C,'─────',round(CDP,2),round(AH,2),round(NH,2),round(NL,2),round(AL,2)],
            '說明': ['今日開盤','今日最高','今日最低','今日收盤','',
                     '=(H+L+2C)/4','=CDP+(H-L)','=2×CDP-L','=2×CDP-H','=CDP-(H-L)']
        })
        summary.to_excel(writer, sheet_name='CDP指標說明', index=False)
        # Sheet4 Strategy
        strategy = pd.DataFrame({
            '情境': ['突破 NH','跌破 NH','突破 NL','跌破 NL','突破 AH','跌破 AL','CDP上方整理','CDP下方整理'],
            '操作建議': ['做多，目標AH','停利/做空，目標NL','逢低買進，目標CDP','做空，目標AL',
                        '強勢追多','強勢追空','回測CDP買進','反彈CDP賣出'],
            '風險提示': ['量需放大','小心假突破','確認不再破NL','量縮留意','超買注意回調','超賣注意反彈','縮量謹慎','縮量謹慎']
        })
        strategy.to_excel(writer, sheet_name='交易策略說明', index=False)

    buf.seek(0)
    wb = load_workbook(buf)

    # 美化 Sheet1
    ws1 = wb['逐筆明細']
    style_header(ws1, 8)
    for row in ws1.iter_rows(min_row=2):
        zone_val = row[7].value
        for cell in row:
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        if zone_val in zone_bg:
            row[7].fill = PatternFill('solid', start_color=zone_bg[zone_val])
            row[7].font = Font(name='Arial', size=9, bold=True, color=zone_fc.get(zone_val,'000000'))
    ws1.freeze_panes = 'A3'
    ws1.auto_filter.ref = f"A2:{get_column_letter(8)}2"
    auto_width(ws1)
    add_title(ws1, f'逐筆成交明細 + CDP區間  |  CDP:{CDP:.2f}  AH:{AH:.2f}  NH:{NH:.2f}  NL:{NL:.2f}  AL:{AL:.2f}', 8)

    # 美化 Sheet2
    ws2 = wb['30分K線CDP分析']
    style_header(ws2, len(col_order))
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        zone_val = row[11].value
        if zone_val in zone_bg:
            row[11].fill = PatternFill('solid', start_color=zone_bg[zone_val])
            row[11].font = Font(name='Arial', size=9, bold=True, color=zone_fc.get(zone_val,'000000'))
        sig_val = row[12].value
        if sig_val and sig_val != '—':
            row[12].fill = PatternFill('solid', start_color='FFF0AA')
            row[12].font = Font(name='Arial', size=9, bold=True, color='8B0000')
    ws2.freeze_panes = 'A3'
    auto_width(ws2)
    add_title(ws2, f'30分K線滾動CDP分析  |  O={O} H={H} L={L} C={C}  Vol={total_vol:,.0f}張', len(col_order))

    # 美化 Sheet3
    ws3 = wb['CDP指標說明']
    style_header(ws3, 3)
    cdp_colors = {
        'CDP（核心軸線）': 'D6E4F0', 'AH（最高目標）': 'FFD7D7',
        'NH（偏多壓力）': 'FFEAD7', 'NL（偏空支撐）': 'D7F5D7', 'AL（最低目標）': 'E8D6F0'
    }
    for row in ws3.iter_rows(min_row=2):
        label = str(row[0].value)
        bg = cdp_colors.get(label, 'FFFFFF')
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill('solid', start_color=bg)
            cell.border = border
    auto_width(ws3)
    add_title(ws3, 'CDP 指標公式說明與數值', 3)

    # 美化 Sheet4
    ws4 = wb['交易策略說明']
    style_header(ws4, 3)
    for i, row in enumerate(ws4.iter_rows(min_row=2)):
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F8FF'
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.border = border
    auto_width(ws4)
    add_title(ws4, 'CDP 交易策略說明', 3)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.read()

# ── 主要分析函式 ──────────────────────────────────────────────────────────────
def analyze(stock_id: str, prev_close: float):
    stock_id = stock_id.strip()
    if not stock_id.isdigit():
        return "❌ 請輸入正確的股票代號（純數字）", None, None, None, None

    try:
        status = f"⏳ 正在爬取 {stock_id} 的逐筆資料，請稍候..."
        yield status, None, None, None, None

        df = fetch_pchome(stock_id)

        # OHLCV
        O = df['成交價'].iloc[0]
        H = df['成交價'].max()
        L = df['成交價'].min()
        C = df['成交價'].iloc[-1]
        total_vol = df['分量(張)'].sum()

        # 整體 CDP
        CDP, AH, NH, NL, AL = calc_cdp(H, L, C)
        df['CDP區間'] = df['成交價'].apply(lambda p: zone_label(p, CDP, AH, NH, NL, AL))

        # 30分K
        df_30 = build_30min_kdp(df, prev_close)

        # CDP 摘要文字
        summary_text = f"""
## 📊 股票代號：{stock_id}

| 指標 | 數值 |
|------|------|
| 開盤(O) | {O} |
| 最高(H) | {H} |
| 最低(L) | {L} |
| 收盤(C) | {C} |
| 總成交量 | {total_vol:,.0f} 張 |

---

| CDP 指標 | 數值 | 說明 |
|----------|------|------|
| **AH** | **{AH:.2f}** | 最高目標（強勢突破） |
| **NH** | **{NH:.2f}** | 偏多壓力（做多參考） |
| **CDP** | **{CDP:.2f}** | 核心軸線 |
| **NL** | **{NL:.2f}** | 偏空支撐（做空參考） |
| **AL** | **{AL:.2f}** | 最低目標（弱勢破底） |

---
> 目前收盤區間：**{zone_label(C, CDP, AH, NH, NL, AL)}**
"""

        # 輸出 Excel
        excel_bytes = build_excel(df, df_30, (CDP, AH, NH, NL, AL), (O, H, L, C, total_vol))
        excel_path = f"/tmp/CDP分析_{stock_id}.xlsx"
        with open(excel_path, 'wb') as f:
            f.write(excel_bytes)

        # 顯示 DataFrame（30分K）
        col_show = ['時間','開盤','最高','最低','收盤','成交量(張)','CDP','AH','NH','NL','AL','區間','進場信號']
        df_show = df_30[col_show].copy()
        for c in ['CDP','AH','NH','NL','AL']:
            df_show[c] = df_show[c].round(2)

        # 逐筆明細（最後20筆）
        df_detail = df[['時間','買價','賣價','成交價','漲跌','分量(張)','累計量(張)','CDP區間']].tail(20)

        yield summary_text, df_show, df_detail, excel_path, f"✅ {stock_id} 分析完成！共 {len(df)} 筆逐筆資料"

    except Exception as e:
        yield f"❌ 發生錯誤：{e}", None, None, None, None

# ── Gradio UI ────────────────────────────────────────────────────────────────
with gr.Blocks(
    title="📈 價量明細 + CDP 交易指標分析",
    theme=gr.themes.Soft(primary_hue="blue"),
    css="""
        .header-box { background: linear-gradient(135deg,#1F3A6E,#2E6DA4); border-radius:10px; padding:20px; margin-bottom:15px; }
        .header-box h1 { color:white; margin:0; font-size:1.5em; }
        .header-box p  { color:#B8D4F0; margin:5px 0 0; font-size:.9em; }
    """
) as demo:

    gr.HTML("""
    <div class="header-box">
        <h1>📈 價量明細 + CDP 交易指標分析系統</h1>
        <p>爬取 PChome 股市逐筆成交資料，自動計算 CDP 關鍵指標並產出 Excel 報告</p>
    </div>
    """)

    with gr.Row():
        with gr.Column(scale=1):
            stock_input = gr.Textbox(
                label="📌 股票代號",
                placeholder="例如：6197、2317、2330",
                value="6197"
            )
            prev_close_input = gr.Number(
                label="昨日收盤價（用於第一根K線CDP估算）",
                value=197.0
            )
            analyze_btn = gr.Button("🔍 開始分析", variant="primary", size="lg")

        with gr.Column(scale=2):
            status_out = gr.Textbox(label="狀態", interactive=False)
            summary_out = gr.Markdown(label="CDP 摘要")

    with gr.Tabs():
        with gr.Tab("📊 30分K線 + CDP"):
            df_30_out = gr.DataFrame(
                label="30分K線 CDP 分析",
                wrap=True,
                interactive=False
            )
        with gr.Tab("📋 逐筆明細（最後20筆）"):
            df_detail_out = gr.DataFrame(
                label="逐筆成交明細",
                wrap=True,
                interactive=False
            )

    with gr.Row():
        excel_out = gr.File(label="📥 下載 Excel 報告（含4個工作表）")

    # 說明
    gr.Markdown("""
    ---
    ### 📖 CDP 指標說明
    | 指標 | 公式 | 意義 |
    |------|------|------|
    | **CDP** | (H+L+2C)/4 | 核心支撐壓力軸 |
    | **AH** | CDP+(H-L) | 最高目標，強勢突破點 |
    | **NH** | 2×CDP-L | 偏多壓力，做空參考 |
    | **NL** | 2×CDP-H | 偏空支撐，做多參考 |
    | **AL** | CDP-(H-L) | 最低目標，弱勢破底點 |

    ### 💡 使用說明
    1. 輸入股票代號（台股4位數）
    2. 輸入昨日收盤價（影響第一根K線的CDP估算）
    3. 點擊「開始分析」，等待約 10-15 秒
    4. 下載 Excel 報告（含逐筆明細、30分K線、CDP說明、交易策略）
    """)

    # 事件綁定
    analyze_btn.click(
        fn=analyze,
        inputs=[stock_input, prev_close_input],
        outputs=[summary_out, df_30_out, df_detail_out, excel_out, status_out],
        show_progress=True
    )

demo.queue().launch(
    server_name="0.0.0.0",
    server_port=None,  # 自動選可用 port，避免 7860 被佔用
    share=True,        # Colab 必須 True 才能產生公開連結
)
